import openpyxl

import conf

"""
read log_sdr_sensor.log file
if first read string is "ERROR - ",
every three row data as a test item result, set a skip_flag as "three loop", skip write excel
every eight keyword save as a test item result,
make up String data, then write specify cell into excel 
write into excel
"""


def write_excel(result_str, cell_row):
    # get excel workbook
    workbook = openpyxl.load_workbook("../" + conf.TestPlanName)
    # get workbook name, return data is list
    # print workbook.sheetnames

    # get valid workobject
    ws = workbook.get_sheet_by_name("Sheet1")

    cell_name = "B" + str(cell_row)
    # modify specify cell value
    ws[cell_name] = result_str

    print cell_name

    # save modify file, parameter is saved file name
    workbook.save("../" + conf.TestPlanName)


# read log_sdr_sensor.log file
with open("log_sdr_sensor.log") as log_file:
    skip_flag = 0
    loop_flag = 1
    result_list = ["Sensor name, ", "sensor number, ", "entity id, ", "sensor type, ", "LC, ", "LNC, ", "UNC, ", "UC"]
    result_str = ""
    cell_row = 1
    while True:
        log_file_data = log_file.readline()
        # read file end, exit loop
        if not log_file_data:
            break
        # if first read string is "ERROR - ",
        # every three row data as a test item result, set a skip_flag as "three loop", skip write excel
        if "spec don't have" in log_file_data:
            result_str += result_list[0]
            skip_flag = 2
            continue
        if skip_flag:
            skip_flag -= 1
            continue

        # every eight keyword save as a test item result, make up String data
        if "ERROR" in log_file_data:
            result_str += result_list[loop_flag-1]
        if loop_flag == 8:
            # write specify cell into excel
            write_excel(result_str, cell_row)
            cell_row += 1
            result_str = ""
            loop_flag = 0
        loop_flag += 1




